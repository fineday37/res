import os
import sys
import shutil
from config import setting
import openpyxl
from openpyxl.styles import Font, Alignment
import configparser
sys.path.append(os.path.dirname(os.path.dirname(__file__)))


cf = configparser.ConfigParser()
cf.read(setting.SOURCE_FILE)


class WriteExcel:
    def __init__(self, filename):
        self.filename = filename
        if not os.path.exists(self.filename):
            shutil.copyfile(setting.SOURCE_FILE, setting.TARGET_FILE)
        self.wb = openpyxl.load_workbook(self.filename)
        # 默认打开第一页
        self.ws = self.wb.active

    def write_data(self, row_n, value):
        font_GREEN = Font(name="宋体", color="GREEN", bold=True)
        font_RED = Font(name="宋体", color="RED", bold=True)
        font1 = Font(name="宋体", color="DARKYELLOW", bold=True)
        align = Alignment(horizontal="center", vertical="center")
        L_n = "L" + str(row_n)
        M_n = "M" + str(row_n)
        if value == "PASS":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_GREEN
        if value == "FALSE":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_RED
        self.ws[L_n].Aligment = align
        self.ws[M_n].Font = font1
        self.ws[M_n].ALIGMENT = align
        self.wb.save(self.filename)
